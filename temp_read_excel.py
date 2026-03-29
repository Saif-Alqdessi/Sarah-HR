import openpyxl, csv, sys, os

path = os.path.join(os.path.dirname(__file__), 'نموذج طلب التوظيف -شركة قبلان للصناعات الغذائية (3).xlsx')
try:
    wb = openpyxl.load_workbook(path, data_only=True)
except Exception as e:
    print(f"ERROR loading: {e}")
    sys.exit(1)

print("SHEETS:", wb.sheetnames)
print()

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*60}")
    print(f"SHEET: {sname}  (rows={ws.max_row}, cols={ws.max_column})")
    print('='*60)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(80, ws.max_row), values_only=True), 1):
        filtered = [v for v in row if v is not None]
        if filtered:
            print(f"  R{i}: {list(row)}")
